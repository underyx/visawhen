"""Build forms.json: quarterly receipts, approvals, denials and pending
applications for every USCIS form, nationwide and (for the forms USCIS breaks
down that way) per field office.

USCIS publishes two kinds of quarterly reports on its Immigration and
Citizenship Data page, as CSV, PDF or (since FY2024) XLSX:

* "All USCIS Application and Petition Form Types": for every form (and for
  some, per category, e.g. I-130 immediate relatives vs. other relatives) the
  forms received, approved and denied in the quarter, the ones pending at its
  end and, since FY2022, USCIS's median processing time. Since FY2014.
* Per-office reports for N-400 (since FY2014), I-485 (since FY2014) and I-130
  (since FY2020): the same four counts per field office or service center.

Sources, in order:

1. www.uscis.gov directly. It sits behind Akamai, which answers HTTP 403 to
   non-browser clients most of the time.
2. The Wayback Machine, whose crawler *is* let through: the CDX index for the
   reports it already has, and Save Page Now for the listing page and for
   reports it has not captured yet, so a freshly published quarter still shows
   up here within a day.

Downloaded reports are cached in reports/ (gitignored, cached between
workflow runs); the parsed dataset is written to forms.json.

Exit codes: 0 when the dataset was (re)built; 1 when a report cannot be
fetched from any source or no longer parses (a layout change), so that the
workflow fails loudly instead of silently publishing a hole in the history.

`--offline` skips discovery and rebuilds the dataset from the cached reports
alone, for iterating on the parsers.
"""

from __future__ import annotations

import csv
import io
import json
import re
import sys
import time
from collections import defaultdict
from collections.abc import Callable
from dataclasses import dataclass
from datetime import date
from datetime import timedelta
from itertools import pairwise
from pathlib import Path
from typing import Any
from urllib.parse import quote

import pdfplumber
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

HERE = Path(__file__).parent
REPORTS_DIR = HERE / "reports"
OUTPUT_PATH = HERE / "forms.json"

USCIS = "https://www.uscis.gov"
LISTING_URL = f"{USCIS}/tools/reports-and-studies/immigration-and-citizenship-data"
WAYBACK = "https://web.archive.org"
CDX_URL = f"{WAYBACK}/cdx/search/cdx"
DOCUMENT_DIRS = (
    "www.uscis.gov/sites/default/files/document/data/",
    "www.uscis.gov/sites/default/files/document/reports/",
)

TIMEOUT = 60
# (connect, read): CDX queries usually answer in 5-15 s but occasionally hang
CDX_TIMEOUT = (10, 90)
SAVE_TIMEOUT = 180
WAYBACK_ATTEMPTS = 6

session = requests.Session()
session.headers.update(
    {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36"
        ),
        "Accept-Language": "en-US,en;q=0.9",
    }
)


@dataclass(frozen=True)
class ReportFamily:
    """One series of quarterly reports."""

    # Which form's per-office reports these are, or None for the all-forms report
    form: str | None
    # Prefixes of the report file names (the CDX index matches them case-insensitively)
    file_prefixes: tuple[str, ...]
    # Full-text search on the listing page that finds the family's newest reports
    listing_query: str
    # Below this many rows, a report is considered misparsed
    min_rows: int
    # Per-office reports: the number of (Received, Approved, Denied, Pending)
    # column groups, the last one being the total of the others
    groups: int | None = None


FAMILIES = {
    "all_forms": ReportFamily(
        None,
        ("all_forms_performancedata", "quarterly_all_forms"),
        "All USCIS Application and Petition Form Types",
        30,
    ),
    # regular, military, total
    "N-400": ReportFamily("N-400", ("n400_perf",), "N-400 Naturalization", 60, 3),
    # immediate relative, other relative, total
    "I-130": ReportFamily("I-130", ("i130_perf",), "I-130 Alien Relative", 20, 3),
    # family, employment, humanitarian, other, total
    "I-485": ReportFamily("I-485", ("i485_perf",), "I-485 Adjustment of Status", 60, 5),
}
# The listing page also links a CSV next to most PDFs and, for FY2013 and
# earlier, reports with other names and layouts; those are not handled.
REPORT_FILENAME = re.compile(
    r"fy(?P<fy>\d{4}|\d{2})_?q(?:tr)?(?P<q>[1-4])(?:_final|_v\d+(?:\.\d+)?)?\.(?P<ext>xlsx|csv|pdf)$",
    re.IGNORECASE,
)
# When USCIS published a quarter in several formats, prefer the most structured one.
FORMAT_PRIORITY = {"xlsx": 0, "csv": 1, "pdf": 2}
# Reports for FY2013 and earlier have different names and layouts
MIN_FISCAL_YEAR = 2014

STATES = {
    "Alabama": "AL",
    "Alaska": "AK",
    "Arizona": "AZ",
    "Arkansas": "AR",
    "California": "CA",
    "Colorado": "CO",
    "Connecticut": "CT",
    "Delaware": "DE",
    "District of Columbia": "DC",
    "Florida": "FL",
    "Georgia": "GA",
    "Hawaii": "HI",
    "Idaho": "ID",
    "Illinois": "IL",
    "Indiana": "IN",
    "Iowa": "IA",
    "Kansas": "KS",
    "Kentucky": "KY",
    "Louisiana": "LA",
    "Maine": "ME",
    "Maryland": "MD",
    "Massachusetts": "MA",
    "Michigan": "MI",
    "Minnesota": "MN",
    "Mississippi": "MS",
    "Missouri": "MO",
    "Montana": "MT",
    "Nebraska": "NE",
    "Nevada": "NV",
    "New Hampshire": "NH",
    "New Jersey": "NJ",
    "New Mexico": "NM",
    "New York": "NY",
    "North Carolina": "NC",
    "North Dakota": "ND",
    "Ohio": "OH",
    "Oklahoma": "OK",
    "Oregon": "OR",
    "Pennsylvania": "PA",
    "Rhode Island": "RI",
    "South Carolina": "SC",
    "South Dakota": "SD",
    "Tennessee": "TN",
    "Texas": "TX",
    "Utah": "UT",
    "Vermont": "VT",
    "Virginia": "VA",
    "Washington": "WA",
    "West Virginia": "WV",
    "Wisconsin": "WI",
    "Wyoming": "WY",
    "American Samoa": "AS",
    "Guam": "GU",
    "Northern Mariana Islands": "MP",
    "Puerto Rico": "PR",
    "U.S. Virgin Islands": "VI",
    "Virgin Islands": "VI",
    "US Virgin Islands": "VI",
    "Washington, D.C.": "DC",
    "Washington DC": "DC",
}
STATES_BY_NORMALIZED_NAME = {
    re.sub(r"[^a-z]", "", name.lower()): name for name in STATES
}
# Office names the reports consistently misspell, and service centers, which
# the I-130 and I-485 reports list by their state's name alone
NAME_FIXES = {
    "OFM": "Fort Myers",
    "CSC": "California Service Center",
    "ESC": "Vermont Service Center",
    "NSC": "Nebraska Service Center",
    "SSC": "Texas Service Center",
    "YSC": "Potomac Service Center",
}
# Immigrant visas (a State Department form) and immigration court
# adjustments: not USCIS adjudications, and only reported since FY2026
SKIPPED_CATEGORIES = {"Supplemental Processing"}

# received, approved, denied and pending
Counts = list[int | None]
# "D" is a count withheld for privacy, "H" one withheld so "D" cannot be derived
VALUE_TOKEN = re.compile(r"^(?:\d{1,3}(?:,\d{3})+|\d+(?:\.\d+)?|D|H|-|N/?A)?$")
NUMBER = re.compile(r"^(?:\d{1,3}(?:,\d{3})+|\d{1,3})$")
# A value, or a piece of one that PDF extraction split off ("1" ",925,486")
NUMBER_FRAGMENT = re.compile(r"^(?:[\d,]+|\d+\.\d+|D|H|-|N/?A)$")
DIGITS = re.compile(r"^[\d,.]+$")
OFFICE_CODE = re.compile(r"^[A-Z]{3}$")
# A "Received" column header (with its footnote digit), as opposed to the
# "Employment-based received at service center" category label above it
RECEIVED_HEADER = re.compile(r"\breceived\s*\d*$", re.IGNORECASE)
# A form number, possibly with a footnote digit or two glued on ("I-6007", "N-400 17")
FORM_KEY = re.compile(r"^(?P<form>[A-Z]{1,4}-\d{2,3}[A-Z]{0,2})\s*\d{0,2}$")
# The office name may itself contain a three-letter word ("Dover AFB DVD"); the
# code is the last one before the counts.
PDF_OFFICE_LINE = re.compile(
    r"^(?P<name>[A-Za-z][A-Za-z .'()-]*)\s+(?P<code>[A-Z]{3})\s+(?P<values>(?:[\d,]+|D|H|-|N/?A)(?:\s+(?:[\d,]+|D|H|-|N/?A)){7,})\s*$"
)
PDF_TOTAL_LINE = re.compile(r"^(?:Grand\s+)?Total\s+(?P<values>.+)$", re.IGNORECASE)
# Reports before FY2017 name the offices without a code
PDF_CODELESS_OFFICE_LINE = re.compile(
    r"^(?P<name>[A-Za-z][A-Za-z .'()-]*?)\s+(?P<values>(?:[\d,]+|D|H|-|N/?A)(?:\s+(?:[\d,]+|D|H|-|N/?A)){7,})\s*$"
)
PDF_FORM_LINE = re.compile(
    r"^(?P<form>[A-Z]{1,4}-\d{2,3}[A-Z]{0,2})(?:\s\d{1,2})?\s+(?P<rest>[A-Za-z(].*)$"
)


@dataclass(frozen=True, order=True)
class Report:
    family: str
    fiscal_year: int
    fiscal_quarter: int
    priority: int
    url: str

    @classmethod
    def from_url(cls, url: str) -> Report | None:
        url = url.split("?")[0]
        filename = url.rsplit("/", 1)[-1].lower()
        family = next(
            (
                name
                for name, spec in FAMILIES.items()
                if filename.startswith(spec.file_prefixes)
            ),
            None,
        )
        match = REPORT_FILENAME.search(filename)
        if family is None or match is None:
            return None
        fiscal_year = int(match["fy"])
        fiscal_year += 2000 if fiscal_year < 100 else 0
        if fiscal_year < MIN_FISCAL_YEAR:
            return None
        return cls(
            family=family,
            fiscal_year=fiscal_year,
            fiscal_quarter=int(match["q"]),
            priority=FORMAT_PRIORITY[match["ext"]],
            url=url,
        )

    @property
    def ext(self) -> str:
        return self.url.rsplit(".", 1)[-1].lower()

    @property
    def cache_path(self) -> Path:
        return (
            REPORTS_DIR
            / f"{self.family.lower()}_fy{self.fiscal_year}q{self.fiscal_quarter}.{self.ext}"
        )

    @property
    def key(self) -> tuple[int, int]:
        return (self.fiscal_year, self.fiscal_quarter)


def calendar_quarter(fiscal_year: int, fiscal_quarter: int) -> tuple[str, date, date]:
    """(key like "2025-Q3", first day, last day) of a fiscal quarter's calendar quarter.

    Fiscal quarter 1 is October-December of the previous calendar year.
    """
    first_month = {1: 10, 2: 1, 3: 4, 4: 7}[fiscal_quarter]
    year = fiscal_year - 1 if fiscal_quarter == 1 else fiscal_year
    start = date(year, first_month, 1)
    end = (start.replace(day=28) + timedelta(days=4 + 31 * 2)).replace(
        day=1
    ) - timedelta(days=1)
    return f"{year}-Q{(first_month - 1) // 3 + 1}", start, end


@dataclass
class OfficeRow:
    state: str | None
    name: str
    code: str | None
    counts: Counts


@dataclass
class OfficeReport:
    offices: list[OfficeRow]
    # Every "Total" row found: the grand total at the top, and in FY2017-FY2020
    # reports also a subtotal at the end of the international offices section.
    totals: list[Counts]

    @property
    def total(self) -> Counts | None:
        """The report's grand total, or None when no trustworthy one was found.

        The grand total is the largest total row; it is dropped when it comes out
        smaller than the field offices it sums (a PDF whose digits split apart).
        """
        received_sum = sum(row.counts[0] or 0 for row in self.offices)
        candidates = [
            values for values in self.totals if (values[0] or 0) >= 0.9 * received_sum
        ]
        return (
            max(candidates, key=lambda values: values[0] or 0) if candidates else None
        )


@dataclass
class FormRow:
    """One line of an all-forms report, for one fiscal quarter."""

    fiscal_year: int
    fiscal_quarter: int
    form: str
    title: str
    category: str | None
    counts: Counts
    processing_time: float | None


# --- fetching -----------------------------------------------------------------


def fetch_wayback(
    url: str,
    params: dict[str, str] | None = None,
    timeout: float | tuple[float, float] = TIMEOUT,
    allow_redirects: bool = True,
) -> requests.Response:
    """GET a web.archive.org URL, retrying on the archive's frequent 429/5xx hiccups and resets."""
    for attempt in range(WAYBACK_ATTEMPTS):
        try:
            r = session.get(
                url, params=params, timeout=timeout, allow_redirects=allow_redirects
            )
            if r.status_code in (200, 302):
                return r
            reason = f"HTTP {r.status_code}"
        except requests.RequestException as e:
            reason = repr(e)
        print(
            f"wayback attempt {attempt + 1}/{WAYBACK_ATTEMPTS} failed ({reason}): {url}"
        )
        time.sleep(5 * 2**attempt)
    raise RuntimeError(
        f"Wayback Machine did not serve {url} after {WAYBACK_ATTEMPTS} attempts"
    )


def wayback_captures(url_prefix: str) -> dict[str, str]:
    """Newest HTTP 200 capture timestamp of every archived URL starting with `url_prefix`."""
    r = fetch_wayback(
        CDX_URL,
        {
            "url": f"{url_prefix}*",
            "output": "json",
            "fl": "timestamp,original",
            "filter": "statuscode:200",
            "collapse": "urlkey",
        },
        timeout=CDX_TIMEOUT,
    )
    rows = r.json()
    captures: dict[str, str] = {}
    for timestamp, original in rows[1:]:
        original = original.split("?")[0]
        if timestamp > captures.get(original, ""):
            captures[original] = timestamp
    return captures


def fetch_capture(timestamp: str, url: str) -> bytes:
    """The unmodified (`id_`) body of one Wayback capture."""
    return fetch_wayback(f"{WAYBACK}/web/{timestamp}id_/{url}").content


def save_page_now(url: str) -> str | None:
    """Ask the Wayback Machine to capture `url` right now; the new capture's timestamp, or None."""
    print(f"asking the Wayback Machine to capture {url}")
    try:
        r = fetch_wayback(
            f"{WAYBACK}/save/{url}", timeout=SAVE_TIMEOUT, allow_redirects=False
        )
    except RuntimeError as e:
        print(f"::warning::{e}")
        return None
    match = re.search(r"/web/(\d{14})", r.headers.get("Location", "") or r.url)
    if match is None:
        print(f"::warning::Save Page Now did not return a capture for {url}")
        return None
    return match[1]


def looks_like(ext: str, content: bytes) -> bool:
    if ext == "xlsx":
        return content.startswith(b"PK")
    if ext == "pdf":
        return content.startswith(b"%PDF")
    text = content[:4000].decode("utf-8", errors="replace").lower()
    return "<html" not in text and "received" in text


def fetch_uscis(url: str, params: dict[str, str] | None = None) -> bytes | None:
    """GET a www.uscis.gov URL directly; None when Akamai blocks us."""
    try:
        r = session.get(url, params=params, timeout=TIMEOUT)
    except requests.RequestException as e:
        print(f"uscis.gov request failed ({e!r}): {url}")
        return None
    if r.status_code != 200:
        print(f"uscis.gov answered HTTP {r.status_code}: {url}")
        return None
    return r.content


def download_report(report: Report, captures: dict[str, str]) -> bytes:
    if report.cache_path.exists():
        return report.cache_path.read_bytes()
    print(f"downloading {report.url}")
    content = fetch_uscis(report.url)
    if content is None or not looks_like(report.ext, content):
        timestamp = captures.get(report.url) or save_page_now(report.url)
        if timestamp is None:
            raise RuntimeError(
                f"{report.url} is blocked and the Wayback Machine has no capture of it"
            )
        content = fetch_capture(timestamp, report.url)
    if not looks_like(report.ext, content):
        raise RuntimeError(f"{report.url} did not come back as a {report.ext} file")
    REPORTS_DIR.mkdir(exist_ok=True)
    report.cache_path.write_bytes(content)
    return content


# --- discovery ----------------------------------------------------------------


def listing_reports(html: bytes) -> set[Report]:
    soup = BeautifulSoup(html, "lxml")
    reports = set()
    for link in soup.find_all("a", href=True):
        href = str(link["href"])
        if href.startswith("/"):
            href = USCIS + href
        report = Report.from_url(href)
        if report is not None:
            reports.add(report)
    return reports


def listing_url(family: ReportFamily) -> str:
    return f"{LISTING_URL}?query={quote(family.listing_query)}&items_per_page=100"


def discover_reports() -> tuple[list[Report], dict[str, str]]:
    """Every known report (one per family and quarter, best format) and the Wayback captures of report files."""
    captures: dict[str, str] = {}
    for directory in DOCUMENT_DIRS:
        for family in FAMILIES.values():
            for prefix in family.file_prefixes:
                captures.update(wayback_captures(f"https://{directory}{prefix}"))
    reports = {
        report for url in captures if (report := Report.from_url(url)) is not None
    }
    print(f"{len(reports)} report files in the Wayback Machine's index")

    for family in FAMILIES.values():
        url = listing_url(family)
        html = fetch_uscis(url)
        if html is None or not listing_reports(html):
            # The listing page is blocked too; have the archive fetch it for us
            # so newly published reports are discovered the same day.
            timestamp = save_page_now(url)
            html = fetch_capture(timestamp, url) if timestamp else None
        listed = listing_reports(html) if html else set()
        print(
            f"{family.listing_query!r}: {len(listed)} report files listed, {len(listed - reports)} new"
        )
        reports |= listed

    best: dict[tuple[str, int, int], Report] = {}
    for report in sorted(reports):
        best.setdefault((report.family, *report.key), report)
    return sorted(best.values()), captures


# --- parsing helpers ----------------------------------------------------------


def parse_value(text: str) -> int | None:
    """One cell: a count, 0 for "-" (represents zero), None for withheld ("D", "H"), "N/A" and blanks."""
    text = text.strip().replace(",", "")
    return int(text) if text.isdigit() else 0 if text == "-" else None


def parse_counts(cells: list[str]) -> Counts | None:
    """Received, approved, denied and pending, or None when these cells are not a data row."""
    cells = [cell.strip() for cell in cells]
    if (
        len(cells) != 4
        or not any(cells)
        or not all(VALUE_TOKEN.match(c) for c in cells)
    ):
        return None
    return [parse_value(cell) for cell in cells]


def parse_months(text: str) -> float | None:
    """A processing time in months; 0 and N/A both mean USCIS did not publish one."""
    try:
        months = float(text.strip().replace(",", ""))
    except ValueError:
        return None
    return months if months > 0 else None


def state_name(text: str) -> str | None:
    return STATES_BY_NORMALIZED_NAME.get(re.sub(r"[^a-z]", "", text.lower()))


def is_total_label(text: str) -> bool:
    return re.sub(r"[^a-z]", "", text.lower()) in (
        "total",
        "grandtotal",
        "totalallforms",
    )


def normalize_dashes(text: str) -> str:
    """Some reports write their zeros with a Unicode hyphen or dash."""
    return re.sub("[\u2010\u2011\u2012\u2013\u2014\u2212]", "-", text)


def grid_of(report: Report, content: bytes) -> list[list[str]]:
    """A CSV or XLSX report as rows of stripped cell strings."""
    if report.ext == "csv":
        text = content.decode("utf-8-sig", errors="replace")
        rows = list(csv.reader(io.StringIO(text)))
    else:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        rows = [
            ["" if cell is None else str(cell) for cell in row]
            for row in workbook.worksheets[0].iter_rows(values_only=True)
        ]
    return [[normalize_dashes(" ".join(cell.split())) for cell in row] for row in rows]


def pdf_lines(content: bytes) -> list[str]:
    """The text lines of a PDF, with the numbers its text layer split apart ("1" ",925,486") joined back.

    The fragments of one number touch (no horizontal gap between them), while
    words and columns are at least a couple of points apart.
    """

    def join(words: list[dict[str, Any]]) -> str:
        words = sorted(words, key=lambda word: word["x0"])
        text = words[0]["text"]
        for previous, word in pairwise(words):
            touching = word["x0"] - previous["x1"] < 1
            numeric = DIGITS.match(previous["text"]) and DIGITS.match(word["text"])
            text += ("" if touching and numeric else " ") + word["text"]
        return text

    lines: list[str] = []
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            line: list[dict[str, Any]] = []
            for word in sorted(page.extract_words(), key=lambda word: word["top"]):
                if line and word["top"] - line[0]["top"] > 3:
                    lines.append(join(line))
                    line = []
                line.append(word)
            if line:
                lines.append(join(line))
    return [normalize_dashes(line.strip()) for line in lines]


def parse_addend(text: str) -> int | tuple[int, int]:
    """A value as a number, or as the (low, high) range a withheld/unavailable one may be."""
    value = parse_value(text)
    return value if value is not None else (0, 9)


def sums_to(addends: list[str], total: str) -> bool:
    """Whether `addends` can add up to `total`, given that withheld ("D") values are small."""
    if parse_value(total) is None:
        return True
    parsed = [parse_addend(a) for a in addends]
    low = sum(v if isinstance(v, int) else v[0] for v in parsed)
    high = sum(v if isinstance(v, int) else v[1] for v in parsed)
    return low <= (parse_value(total) or 0) <= high


def groups_sum_to_total(values: list[str], groups: int) -> bool:
    """A per-office row is consistent when, per column, the category groups add up to the total group."""
    return all(
        sums_to(
            [values[4 * g + c] for g in range(groups - 1)], values[4 * (groups - 1) + c]
        )
        for c in range(4)
    )


def rejoin_numbers(
    tokens: list[str],
    count: int,
    consistent: Callable[[list[str]], bool] = lambda values: True,
) -> list[str] | None:
    """Undo the digit splitting of PDF text ("1 ,925,486", "3 3,077") given how many values there must be.

    Adjacent tokens are merged wherever that is the only way to end up with
    `count` well-formed values that pass the `consistent` check (the report's
    own arithmetic); None when there is no way or more than one.
    """

    def solutions(tokens: list[str], count: int) -> list[list[str]]:
        if count == 0:
            return [[]] if not tokens else []
        if not tokens or len(tokens) < count:
            return []
        found: list[list[str]] = []
        for take in range(1, len(tokens) - count + 2):
            value = "".join(tokens[:take])
            if (
                VALUE_TOKEN.match(value)
                and value
                and not (take > 1 and not NUMBER.match(value))
            ):
                found.extend(
                    [value, *rest] for rest in solutions(tokens[take:], count - 1)
                )
                if len(found) > 500:
                    break
        return found

    if len(tokens) == count and consistent(tokens):
        return tokens
    found = [values for values in solutions(tokens, count) if consistent(values)]
    return found[0] if len(found) == 1 else None


# --- per-office reports -------------------------------------------------------


def parse_office_grid(grid: list[list[str]]) -> OfficeReport:
    """Parse the CSV/XLSX layout: label columns, then per category (Received, Approved, Denied, Pending), the last category being the total."""
    first_column: int | None = None
    total_columns: list[int] | None = None
    for row in grid:
        received = [i for i, cell in enumerate(row) if RECEIVED_HEADER.search(cell)]
        if len(received) >= 2:
            first_column = received[0]
            total_columns = [received[-1] + k for k in range(4)]
            break
    if first_column is None or total_columns is None:
        raise ValueError("no header row with 'Received' columns")

    offices: list[OfficeRow] = []
    totals: list[Counts] = []
    state: str | None = None
    for row in grid:
        cells = row + [""] * (max(total_columns) + 1 - len(row))
        labels = cells[:first_column]
        code = next((c for c in labels if OFFICE_CODE.match(c)), None)
        names = [c for c in labels if c and c != code]
        if not names:
            continue
        name = names[-1]
        counts = parse_counts([cells[i] for i in total_columns])
        if is_total_label(name) and counts is not None:
            totals.append(counts)
        elif counts is None:
            # A heading: a state, a country in the international offices
            # section, "Service Centers", or a label like "Field Office by
            # State". Only checked on rows without counts, as the Washington
            # field office (in DC) shares its name with the state.
            if labels[0]:
                state = state_name(name)
            continue
        elif code is not None:
            offices.append(OfficeRow(state, name, code, counts))
        elif labels[0] or state is None:
            # No office code and either the name sits in the state column or it
            # is under a country heading: an international office (code "N/A"
            # in later reports) or a stray line, not a field office.
            if labels[0] and state_name(name) is not None:
                state = state_name(name)
            continue
        else:
            offices.append(OfficeRow(state, name, code, counts))
    return OfficeReport(offices, totals)


def parse_office_pdf(content: bytes, groups: int) -> OfficeReport:
    """Parse the PDF layout from its text layer: one office per line, followed by its counts per category."""
    offices: list[OfficeRow] = []
    totals: list[Counts] = []
    state: str | None = None
    count = 4 * groups

    def values_of(text: str) -> Counts | None:
        # Digits often come out split ("1 94,819"); the per-category counts
        # must add up to the total ones, which settles most of the ambiguity.
        values = rejoin_numbers(
            text.split(), count, lambda values: groups_sum_to_total(values, groups)
        )
        return parse_counts(values[-4:]) if values is not None else None

    for line in pdf_lines(content):
        if (known_state := state_name(line)) is not None:
            state = known_state
        elif (match := PDF_TOTAL_LINE.match(line)) is not None:
            if (counts := values_of(match["values"])) is not None:
                totals.append(counts)
        elif (match := PDF_OFFICE_LINE.match(line)) is not None:
            if (counts := values_of(match["values"])) is not None:
                offices.append(OfficeRow(state, match["name"], match["code"], counts))
            else:
                print(f"::warning::cannot read the numbers of {line!r}")
        elif (
            state is not None
            and (match := PDF_CODELESS_OFFICE_LINE.match(line)) is not None
            and (counts := values_of(match["values"])) is not None
        ):
            offices.append(OfficeRow(state, match["name"], None, counts))
        elif re.fullmatch(r"[A-Za-z][A-Za-z .'&-]+", line):
            # a heading that is not a state: a country in the international
            # offices section, or "Service Centers"
            state = None
    return OfficeReport(offices, totals)


def parse_office_report(report: Report, content: bytes) -> OfficeReport:
    family = FAMILIES[report.family]
    assert family.groups is not None
    parsed = (
        parse_office_pdf(content, family.groups)
        if report.ext == "pdf"
        else parse_office_grid(grid_of(report, content))
    )
    if len(parsed.offices) < FAMILIES[report.family].min_rows:
        raise ValueError(
            f"only {len(parsed.offices)} offices parsed from {report.url}; the layout must have changed"
        )
    return parsed


# --- the all-forms report -----------------------------------------------------


@dataclass
class ColumnGroup:
    fiscal_quarter: int
    received: int
    approved: int
    denied: int
    pending: int
    processing_time: int | None


def all_forms_groups(
    grid: list[list[str]], report: Report
) -> tuple[list[ColumnGroup], int]:
    """The column groups of an all-forms CSV/XLSX and the index of the first data row.

    Until FY2019 Q2 a report had one group per quarter of the fiscal year so far
    (and a year-to-date group); since then it has the quarter and year-to-date.
    """
    # the column header row always has "Pending"; the rows above it hold the
    # group labels ("2nd Quarter") and split words ("Forms" / "Received")
    header_end = next(
        i for i, row in enumerate(grid) if any("pending" in c.lower() for c in row)
    )
    header_rows = grid[max(0, header_end - 3) : header_end + 1]
    width = max(len(row) for row in header_rows)
    labels = [
        " ".join(row[i].lower() for row in header_rows if i < len(row) and row[i])
        for i in range(width)
    ]
    starts = [i for i, label in enumerate(labels) if RECEIVED_HEADER.search(label)]
    # the cumulative layout has a group per quarter plus year-to-date, the
    # single-quarter one just the quarter and year-to-date
    cumulative = len(starts) >= 3
    groups = []
    for quarter, start in enumerate(starts[:4] if cumulative else starts[:1], 1):
        fiscal_quarter = quarter if cumulative else report.fiscal_quarter
        if fiscal_quarter > report.fiscal_quarter:
            break
        pending = next(i for i in range(start + 3, start + 6) if "pending" in labels[i])
        processing_time = next(
            (
                i
                for i in range(start + 3, start + 7)
                if i < width and "time" in labels[i]
            ),
            None,
        )
        groups.append(
            ColumnGroup(
                fiscal_quarter, start, start + 1, start + 2, pending, processing_time
            )
        )
    return groups, header_end + 1


def parse_all_forms_grid(report: Report, grid: list[list[str]]) -> list[FormRow]:
    groups, first_data_row = all_forms_groups(grid, report)
    first_column = groups[0].received
    rows: list[FormRow] = []
    category: str | None = None
    for row in grid[first_data_row:]:
        cells = row + [""] * (max(g.pending for g in groups) + 3 - len(row))
        labels = [c for c in cells[:first_column] if c]
        if not labels:
            continue
        # a form row is "I-130, title" or "Family Based, I-130, title"; a
        # category heading is a lone label without counts
        form_index, form_match = next(
            ((i, m) for i, c in enumerate(labels) if (m := FORM_KEY.match(c))),
            (None, None),
        )
        has_values = any(cells[first_column:])
        if form_match is None or form_index is None:
            if not has_values and not is_total_label(labels[0]):
                category = labels[0]
            continue
        if form_index > 0:
            category = labels[0]
        if category in SKIPPED_CATEGORIES:
            continue
        title = labels[form_index + 1] if len(labels) > form_index + 1 else ""
        for group in groups:
            counts = parse_counts(
                [
                    cells[group.received],
                    cells[group.approved],
                    cells[group.denied],
                    cells[group.pending],
                ]
            )
            if counts is None:
                continue
            rows.append(
                FormRow(
                    report.fiscal_year,
                    group.fiscal_quarter,
                    form_match["form"],
                    title,
                    category,
                    counts,
                    (
                        parse_months(cells[group.processing_time])
                        if group.processing_time is not None
                        else None
                    ),
                )
            )
    return rows


def parse_all_forms_pdf(report: Report, content: bytes) -> list[FormRow]:
    """The PDF layouts (FY2019 Q3 - FY2021 Q3), digits often split by spaces.

    Until FY2020 a report had cumulative fiscal-year columns (four per quarter
    so far, then year-to-date received and approved); from FY2021 it has the
    quarter (received, approved, denied, completions, pending, processing
    time) and, after the first quarter, the year-to-date (the same without
    processing time).
    """
    lines = pdf_lines(content)
    cumulative = not any("processing time" in line.lower() for line in lines[:8])
    quarters = report.fiscal_quarter
    value_count = 4 * quarters + 2 if cumulative else 6 if quarters == 1 else 11

    def consistent(values: list[str]) -> bool:
        if cumulative:
            # the year-to-date received and approved are the sums of the quarters'
            return all(
                sums_to(
                    [values[4 * q + c] for q in range(quarters)],
                    values[4 * quarters + c],
                )
                for c in range(2)
            )
        received, approved, denied, completions = (parse_value(v) for v in values[:4])
        if None not in (approved, denied, completions) and completions < approved + denied:  # type: ignore[operator]
            return False
        if quarters == 1:
            return True
        # the year-to-date received is at least the quarter's
        year_to_date = parse_value(values[6])
        return received is None or year_to_date is None or year_to_date >= received

    rows: list[FormRow] = []
    category: str | None = None
    for line in lines:
        match = PDF_FORM_LINE.match(line)
        if match is None:
            if re.fullmatch(
                r"[A-Za-z][A-Za-z &/]+", line
            ) and not line.lower().startswith("total"):
                category = line
            continue
        if category in SKIPPED_CATEGORIES:
            continue
        words = match["rest"].split()
        # the title is everything up to the first run of value-like tokens that reaches the end
        split = len(words)
        while split > 0 and NUMBER_FRAGMENT.match(words[split - 1]):
            split -= 1
        title = " ".join(words[:split])
        tokens = words[split:]
        values = rejoin_numbers(tokens, value_count, consistent)
        if values is None and not cumulative and len(tokens) == value_count - 1:
            # a blank processing time
            values = rejoin_numbers(
                [*tokens[:5], "N/A", *tokens[5:]], value_count, consistent
            )
        if values is None:
            print(
                f"::warning::FY{report.fiscal_year} Q{report.fiscal_quarter}: cannot read the numbers of {match['form']} {title!r}: {' '.join(words[split:])}"
            )
            continue
        if not cumulative:
            counts = parse_counts([values[0], values[1], values[2], values[4]])
            if counts is not None:
                rows.append(
                    FormRow(
                        report.fiscal_year,
                        quarters,
                        match["form"],
                        title,
                        category,
                        counts,
                        parse_months(values[5]),
                    )
                )
            continue
        for quarter in range(1, quarters + 1):
            counts = parse_counts(values[4 * (quarter - 1) : 4 * quarter])
            if counts is not None:
                rows.append(
                    FormRow(
                        report.fiscal_year,
                        quarter,
                        match["form"],
                        title,
                        category,
                        counts,
                        None,
                    )
                )
    return rows


def parse_all_forms_report(report: Report, content: bytes) -> list[FormRow]:
    rows = (
        parse_all_forms_pdf(report, content)
        if report.ext == "pdf"
        else parse_all_forms_grid(report, grid_of(report, content))
    )
    current = [row for row in rows if row.fiscal_quarter == report.fiscal_quarter]
    if len(current) < FAMILIES["all_forms"].min_rows:
        raise ValueError(
            f"only {len(current)} form rows parsed from {report.url}; the layout must have changed"
        )
    return rows


# --- assembly -----------------------------------------------------------------


def normalize_name(name: str) -> str:
    name = re.sub(r"[^a-z ]", "", name.lower())
    name = re.sub(r"\b(saint|st)\b", "st", name)
    name = re.sub(r"\b(fort|ft)\b", "ft", name)
    return re.sub(r"\s+", " ", name).strip()


def slugify(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")


def counts_dict(counts: Counts) -> dict[str, int | None]:
    received, approved, denied, pending = counts
    return {
        "received": received,
        "approved": approved,
        "denied": denied,
        "pending": pending,
    }


def sum_counts(rows: list[Counts]) -> dict[str, int | None]:
    """Per field, the sum of the rows that have it, or None when none does."""
    result: dict[str, int | None] = {}
    for i, key in enumerate(("received", "approved", "denied", "pending")):
        values: list[int] = [v for row in rows if (v := row[i]) is not None]
        result[key] = sum(values) if values else None
    return result


def base_title(title: str) -> str:
    return re.sub(r"\s*\([^()]*(?:\([^()]*\)[^()]*)*\)\s*$", "", title) or title


def build_offices(
    reports: list[tuple[Report, OfficeReport]],
) -> tuple[list[dict[str, Any]], dict[str, dict[str, int | None]], dict[str, str]]:
    """One form's office pages, its per-quarter totals and the report each quarter came from."""
    # Reports before FY2017 name the offices without their codes; resolve those
    # names against every report that has both, most specific match first.
    codes_by_name: dict[tuple[str | None, str], str] = {}
    for _, parsed in reports:
        for row in parsed.offices:
            if row.code is not None:
                codes_by_name.setdefault(
                    (row.state, normalize_name(row.name)), row.code
                )
                codes_by_name.setdefault((None, normalize_name(row.name)), row.code)

    offices: dict[str, dict[str, Any]] = {}
    totals: dict[str, dict[str, int | None]] = {}
    sources: dict[str, str] = {}
    for report, parsed in reports:
        quarter, _, _ = calendar_quarter(*report.key)
        rows: list[OfficeRow] = []
        for row in parsed.offices:
            code = (
                row.code
                or codes_by_name.get((row.state, normalize_name(row.name)))
                or codes_by_name.get((None, normalize_name(row.name)))
            )
            if code is None:
                print(
                    f"::warning::{report.family} {quarter}: no office code known for {row.name!r} ({row.state}); skipped"
                )
                continue
            rows.append(OfficeRow(row.state, row.name, code, row.counts))
            # The newest report a code appears in names it; reports are in chronological order.
            office = offices.setdefault(code, {"code": code, "quarters": {}})
            office["name"] = NAME_FIXES.get(code, row.name)
            office["state"] = row.state
            office["quarters"][quarter] = counts_dict(row.counts)
        totals[quarter] = (
            counts_dict(parsed.total)
            if parsed.total is not None
            else sum_counts([row.counts for row in rows])
        )
        sources[quarter] = report.url

    for office in offices.values():
        state_code = STATES.get(office["state"]) if office["state"] else None
        office["stateCode"] = state_code
        office["slug"] = slugify(f"{office['name']} {state_code or ''}")
    slugs = [office["slug"] for office in offices.values()]
    if len(set(slugs)) != len(slugs):
        raise ValueError(
            f"duplicate office slugs: {sorted(s for s in slugs if slugs.count(s) > 1)}"
        )
    return sorted(offices.values(), key=lambda office: office["code"]), totals, sources


def build_dataset(
    all_forms: list[tuple[Report, list[FormRow]]],
    office_reports: dict[str, list[tuple[Report, OfficeReport]]],
) -> dict[str, Any]:
    # Newer reports restate earlier quarters (the cumulative layout, and
    # post-adjudicative corrections); the newest report wins for each quarter.
    rows_by_quarter: dict[tuple[str, tuple[int, int]], list[FormRow]] = {}
    sources_by_quarter: dict[tuple[str, tuple[int, int]], str] = {}
    for report, rows in all_forms:
        for row in rows:
            key = (row.form, (row.fiscal_year, row.fiscal_quarter))
            if sources_by_quarter.get(key) != report.url:
                rows_by_quarter[key] = []
                sources_by_quarter[key] = report.url
            rows_by_quarter[key].append(row)

    forms: dict[str, dict[str, Any]] = {}
    for (form, fiscal), rows in sorted(
        rows_by_quarter.items(), key=lambda item: item[0][1]
    ):
        quarter, _, _ = calendar_quarter(*fiscal)
        entry = forms.setdefault(
            form, {"form": form, "slug": slugify(form), "quarters": {}, "sources": {}}
        )
        entry["quarters"][quarter] = {
            **sum_counts([row.counts for row in rows]),
            "variants": [
                {
                    "title": row.title,
                    **counts_dict(row.counts),
                    "processingTime": row.processing_time,
                }
                for row in rows
            ],
        }
        entry["sources"][quarter] = sources_by_quarter[(form, fiscal)]
        # the newest report names and categorizes the form
        main = max(rows, key=lambda row: row.counts[0] or 0)
        entry["title"] = base_title(main.title)
        entry["category"] = main.category

    for family, reports in office_reports.items():
        office_form = FAMILIES[family].form
        assert office_form is not None
        offices, totals, sources = build_offices(reports)
        entry = forms.setdefault(
            office_form,
            {
                "form": office_form,
                "slug": slugify(office_form),
                "quarters": {},
                "sources": {},
            },
        )
        entry["offices"] = offices
        entry["officeTotals"] = totals
        entry["officeSources"] = sources
    for entry in forms.values():
        entry.setdefault("offices", [])
        entry.setdefault("officeTotals", {})
        entry.setdefault("officeSources", {})
        entry.setdefault("title", entry["form"])
        entry.setdefault("category", None)

    quarters = sorted(
        {
            q
            for entry in forms.values()
            for q in (*entry["quarters"], *entry["officeTotals"])
        }
    )
    periods = []
    for quarter in quarters:
        year, q = quarter.split("-Q")
        fiscal_year, fiscal_quarter = (
            (int(year) + 1, 1) if q == "4" else (int(year), int(q) + 1)
        )
        quarter_key, start, end = calendar_quarter(fiscal_year, fiscal_quarter)
        assert quarter_key == quarter
        periods.append(
            {
                "quarter": quarter,
                "start": start.isoformat(),
                "end": end.isoformat(),
                "fiscalYear": fiscal_year,
                "fiscalQuarter": fiscal_quarter,
            }
        )
    return {
        "periods": periods,
        "forms": sorted(forms.values(), key=lambda entry: entry["form"]),
    }


def cached_reports() -> list[Report]:
    """The reports in the cache directory, for iterating on the parsers offline."""
    reports = []
    for path in sorted(REPORTS_DIR.iterdir()):
        match = re.fullmatch(
            r"(?P<family>.+)_fy(?P<fy>\d{4})q(?P<q>[1-4])\.(?P<ext>\w+)", path.name
        )
        if match is None:
            continue
        family = next(name for name in FAMILIES if name.lower() == match["family"])
        reports.append(
            Report(
                family,
                int(match["fy"]),
                int(match["q"]),
                FORMAT_PRIORITY[match["ext"]],
                path.name,
            )
        )
    return sorted(reports)


def main() -> int:
    captures: dict[str, str] = {}
    if "--offline" in sys.argv:
        reports = cached_reports()
    else:
        reports, captures = discover_reports()
    all_forms: list[tuple[Report, list[FormRow]]] = []
    office_reports: dict[str, list[tuple[Report, OfficeReport]]] = defaultdict(list)
    for report in reports:
        content = download_report(report, captures)
        if report.family == "all_forms":
            rows = parse_all_forms_report(report, content)
            all_forms.append((report, rows))
            current = sum(row.fiscal_quarter == report.fiscal_quarter for row in rows)
            print(
                f"all forms FY{report.fiscal_year} Q{report.fiscal_quarter} ({report.ext}): {current} form rows"
            )
        else:
            parsed = parse_office_report(report, content)
            office_reports[report.family].append((report, parsed))
            print(
                f"{report.family} FY{report.fiscal_year} Q{report.fiscal_quarter} ({report.ext}): {len(parsed.offices)} offices,"
                f" total {'from report' if parsed.total else 'summed'}"
            )
    dataset = build_dataset(all_forms, office_reports)
    OUTPUT_PATH.write_text(json.dumps(dataset, indent=2, sort_keys=True) + "\n")
    print(
        f"wrote {OUTPUT_PATH}: {len(dataset['forms'])} forms over {len(dataset['periods'])} quarters"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
