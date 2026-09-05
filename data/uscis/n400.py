"""Build n400.json: quarterly N-400 receipts, approvals, denials and pending
applications per USCIS field office.

USCIS publishes a report called "Form N-400, Application for Naturalization, by
Category of Naturalization, Case Status, and USCIS Field Office Location" for
every fiscal quarter since FY2014 Q1 (October 2013), as CSV, PDF or (since
FY2024) XLSX. Each report has, per field office, the applications received,
approved and denied during the quarter and the ones still pending at its end.

Sources, in order:

1. www.uscis.gov directly. It sits behind Akamai, which answers HTTP 403 to
   non-browser clients most of the time.
2. The Wayback Machine, whose crawler *is* let through: the CDX index for the
   reports it already has, and Save Page Now for the listing page and for
   reports it has not captured yet, so a freshly published quarter still shows
   up here within a day.

Downloaded reports are cached in n400_reports/ (gitignored, cached between
workflow runs); the parsed dataset is written to n400.json.

Exit codes: 0 when the dataset was (re)built; 1 when a report cannot be
fetched from any source or no longer parses (a layout change), so that the
workflow fails loudly instead of silently publishing a hole in the history.
"""

from __future__ import annotations

import csv
import io
import json
import re
import sys
import time
from dataclasses import dataclass
from datetime import date
from datetime import timedelta
from pathlib import Path
from typing import Any

import pdfplumber
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

HERE = Path(__file__).parent
REPORTS_DIR = HERE / "n400_reports"
OUTPUT_PATH = HERE / "n400.json"

USCIS = "https://www.uscis.gov"
# The "Naturalization Data" topic of the Immigration and Citizenship Data page
LISTING_URL = f"{USCIS}/tools/reports-and-studies/immigration-and-citizenship-data?topic_id%5B%5D=33692&items_per_page=100"
WAYBACK = "https://web.archive.org"
CDX_URL = f"{WAYBACK}/cdx/search/cdx"
# The CDX index matches these prefixes case-insensitively, so they also cover
# the N400_performancedata_* and N400_Performancedata_* spellings.
CDX_PREFIXES = (
    "www.uscis.gov/sites/default/files/document/data/n400_perf",
    "www.uscis.gov/sites/default/files/document/reports/n400_perf",
)
REPORT_FILENAME = re.compile(
    r"n400_perf\w*?_?fy(?P<fy>\d{4})_?q(?:tr)?(?P<q>[1-4])(?:_v\d+)?\.(?P<ext>xlsx|csv|pdf)$",
    re.IGNORECASE,
)
# When USCIS published a quarter in several formats, prefer the most structured one.
FORMAT_PRIORITY = {"xlsx": 0, "csv": 1, "pdf": 2}

TIMEOUT = 60
# (connect, read): CDX queries usually answer in 5-15 s but occasionally hang
CDX_TIMEOUT = (10, 90)
SAVE_TIMEOUT = 180
WAYBACK_ATTEMPTS = 4

session = requests.Session()
session.headers.update(
    {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36"
        ),
        "Accept-Language": "en-US,en;q=0.9",
    }
)

STATES = {
    "Alabama": "AL",
    "Alaska": "AK",
    "Arizona": "AZ",
    "Arkansas": "AR",
    "California": "CA",
    "Colorado": "CO",
    "Connecticut": "CT",
    "Delaware": "DE",
    "District of Columbia": "DC",
    "Florida": "FL",
    "Georgia": "GA",
    "Hawaii": "HI",
    "Idaho": "ID",
    "Illinois": "IL",
    "Indiana": "IN",
    "Iowa": "IA",
    "Kansas": "KS",
    "Kentucky": "KY",
    "Louisiana": "LA",
    "Maine": "ME",
    "Maryland": "MD",
    "Massachusetts": "MA",
    "Michigan": "MI",
    "Minnesota": "MN",
    "Mississippi": "MS",
    "Missouri": "MO",
    "Montana": "MT",
    "Nebraska": "NE",
    "Nevada": "NV",
    "New Hampshire": "NH",
    "New Jersey": "NJ",
    "New Mexico": "NM",
    "New York": "NY",
    "North Carolina": "NC",
    "North Dakota": "ND",
    "Ohio": "OH",
    "Oklahoma": "OK",
    "Oregon": "OR",
    "Pennsylvania": "PA",
    "Rhode Island": "RI",
    "South Carolina": "SC",
    "South Dakota": "SD",
    "Tennessee": "TN",
    "Texas": "TX",
    "Utah": "UT",
    "Vermont": "VT",
    "Virginia": "VA",
    "Washington": "WA",
    "West Virginia": "WV",
    "Wisconsin": "WI",
    "Wyoming": "WY",
    "American Samoa": "AS",
    "Guam": "GU",
    "Northern Mariana Islands": "MP",
    "Puerto Rico": "PR",
    "U.S. Virgin Islands": "VI",
    "Virgin Islands": "VI",
    "US Virgin Islands": "VI",
    "Washington, D.C.": "DC",
    "Washington DC": "DC",
}
STATES_BY_NORMALIZED_NAME = {
    re.sub(r"[^a-z]", "", name.lower()): name for name in STATES
}
# Office names the reports consistently misspell
NAME_FIXES = {"OFM": "Fort Myers"}

# One report row's twelve numbers: received, approved, denied and pending, for
# non-military naturalization, military naturalization, and both combined.
Values = list[int | None]
VALUE_TOKEN = re.compile(r"^(?:\d{1,3}(?:,\d{3})+|\d+|D|-|N/A)?$")
CODE = re.compile(r"^[A-Z]{3}$")
PDF_OFFICE_LINE = re.compile(
    r"^(?P<name>[A-Za-z][A-Za-z .'()-]*?)\s+(?P<code>[A-Z]{3})\s+(?P<values>\S+(?:\s+\S+){11})\s*$"
)
PDF_TOTAL_LINE = re.compile(r"^(?:Grand\s+)?Total\s+(?P<values>.+)$", re.IGNORECASE)


@dataclass(frozen=True, order=True)
class Report:
    fiscal_year: int
    fiscal_quarter: int
    priority: int
    url: str

    @classmethod
    def from_url(cls, url: str) -> Report | None:
        match = REPORT_FILENAME.search(url.split("?")[0])
        if match is None:
            return None
        return cls(
            fiscal_year=int(match["fy"]),
            fiscal_quarter=int(match["q"]),
            priority=FORMAT_PRIORITY[match["ext"].lower()],
            url=url.split("?")[0],
        )

    @property
    def ext(self) -> str:
        return self.url.rsplit(".", 1)[-1].lower()

    @property
    def cache_path(self) -> Path:
        return REPORTS_DIR / f"fy{self.fiscal_year}q{self.fiscal_quarter}.{self.ext}"

    @property
    def calendar_quarter(self) -> tuple[str, date, date]:
        """(key like "2025-Q3", first day, last day) of the calendar quarter this report covers.

        Fiscal quarter 1 is October-December of the previous calendar year.
        """
        first_month = {1: 10, 2: 1, 3: 4, 4: 7}[self.fiscal_quarter]
        year = self.fiscal_year - 1 if self.fiscal_quarter == 1 else self.fiscal_year
        start = date(year, first_month, 1)
        end = (start.replace(day=28) + timedelta(days=4 + 31 * 2)).replace(
            day=1
        ) - timedelta(days=1)
        return f"{year}-Q{(first_month - 1) // 3 + 1}", start, end


@dataclass
class OfficeRow:
    state: str | None
    name: str
    code: str | None
    counts: Values


@dataclass
class ParsedReport:
    offices: list[OfficeRow]
    # Every "Total" row found: the grand total at the top, and in FY2017-FY2020
    # reports also a subtotal at the end of the international offices section.
    totals: list[Values]

    @property
    def total(self) -> Values | None:
        """The report's grand total, or None when no trustworthy one was found.

        The grand total is the largest total row; it is dropped when it comes out
        smaller than the field offices it sums (a PDF whose digits split apart).
        """
        received_sum = sum(row.counts[8] or 0 for row in self.offices)
        candidates = [
            values for values in self.totals if (values[8] or 0) >= 0.9 * received_sum
        ]
        return (
            max(candidates, key=lambda values: values[8] or 0) if candidates else None
        )


# --- fetching -----------------------------------------------------------------


def fetch_wayback(
    url: str,
    params: dict[str, str] | None = None,
    timeout: float | tuple[float, float] = TIMEOUT,
    allow_redirects: bool = True,
) -> requests.Response:
    """GET a web.archive.org URL, retrying on the archive's frequent 429/5xx hiccups and resets."""
    for attempt in range(WAYBACK_ATTEMPTS):
        try:
            r = session.get(
                url, params=params, timeout=timeout, allow_redirects=allow_redirects
            )
            if r.status_code in (200, 302):
                return r
            reason = f"HTTP {r.status_code}"
        except requests.RequestException as e:
            reason = repr(e)
        print(
            f"wayback attempt {attempt + 1}/{WAYBACK_ATTEMPTS} failed ({reason}): {url}"
        )
        time.sleep(5 * 2**attempt)
    raise RuntimeError(
        f"Wayback Machine did not serve {url} after {WAYBACK_ATTEMPTS} attempts"
    )


def wayback_captures(url_prefix: str) -> dict[str, str]:
    """Newest HTTP 200 capture timestamp of every archived URL starting with `url_prefix`."""
    r = fetch_wayback(
        CDX_URL,
        {
            "url": f"{url_prefix}*",
            "output": "json",
            "fl": "timestamp,original",
            "filter": "statuscode:200",
            "collapse": "urlkey",
        },
        timeout=CDX_TIMEOUT,
    )
    rows = r.json()
    captures: dict[str, str] = {}
    for timestamp, original in rows[1:]:
        original = original.split("?")[0]
        if timestamp > captures.get(original, ""):
            captures[original] = timestamp
    return captures


def fetch_capture(timestamp: str, url: str) -> bytes:
    """The unmodified (`id_`) body of one Wayback capture."""
    return fetch_wayback(f"{WAYBACK}/web/{timestamp}id_/{url}").content


def save_page_now(url: str) -> str | None:
    """Ask the Wayback Machine to capture `url` right now; the new capture's timestamp, or None."""
    print(f"asking the Wayback Machine to capture {url}")
    try:
        r = fetch_wayback(
            f"{WAYBACK}/save/{url}", timeout=SAVE_TIMEOUT, allow_redirects=False
        )
    except RuntimeError as e:
        print(f"::warning::{e}")
        return None
    match = re.search(r"/web/(\d{14})", r.headers.get("Location", "") or r.url)
    if match is None:
        print(f"::warning::Save Page Now did not return a capture for {url}")
        return None
    return match[1]


def looks_like(ext: str, content: bytes) -> bool:
    if ext == "xlsx":
        return content.startswith(b"PK")
    if ext == "pdf":
        return content.startswith(b"%PDF")
    text = content[:4000].decode("utf-8", errors="replace").lower()
    return "<html" not in text and "naturalization" in text


def fetch_uscis(url: str) -> bytes | None:
    """GET a www.uscis.gov URL directly; None when Akamai blocks us."""
    try:
        r = session.get(url, timeout=TIMEOUT)
    except requests.RequestException as e:
        print(f"uscis.gov request failed ({e!r}): {url}")
        return None
    if r.status_code != 200:
        print(f"uscis.gov answered HTTP {r.status_code}: {url}")
        return None
    return r.content


def download_report(report: Report, captures: dict[str, str]) -> bytes:
    if report.cache_path.exists():
        return report.cache_path.read_bytes()
    print(f"downloading {report.url}")
    content = fetch_uscis(report.url)
    if content is None or not looks_like(report.ext, content):
        timestamp = captures.get(report.url) or save_page_now(report.url)
        if timestamp is None:
            raise RuntimeError(
                f"{report.url} is blocked and the Wayback Machine has no capture of it"
            )
        content = fetch_capture(timestamp, report.url)
    if not looks_like(report.ext, content):
        raise RuntimeError(f"{report.url} did not come back as a {report.ext} file")
    REPORTS_DIR.mkdir(exist_ok=True)
    report.cache_path.write_bytes(content)
    return content


# --- discovery ----------------------------------------------------------------


def listing_reports(html: bytes) -> set[Report]:
    soup = BeautifulSoup(html, "lxml")
    reports = set()
    for link in soup.find_all("a", href=True):
        href = str(link["href"])
        if href.startswith("/"):
            href = USCIS + href
        report = Report.from_url(href)
        if report is not None:
            reports.add(report)
    return reports


def discover_reports() -> tuple[list[Report], dict[str, str]]:
    """Every known report (one per quarter, best format) and the Wayback captures of report files."""
    captures: dict[str, str] = {}
    for prefix in CDX_PREFIXES:
        captures.update(wayback_captures(f"https://{prefix}"))
    reports = {
        report for url in captures if (report := Report.from_url(url)) is not None
    }
    print(f"{len(reports)} report files in the Wayback Machine's index")

    html = fetch_uscis(LISTING_URL)
    if html is None or not listing_reports(html):
        # The listing page is blocked too; have the archive fetch it for us
        # so newly published reports are discovered the same day.
        timestamp = save_page_now(LISTING_URL)
        html = fetch_capture(timestamp, LISTING_URL) if timestamp else None
    listed = listing_reports(html) if html else set()
    print(
        f"{len(listed)} report files on the listing page, {len(listed - reports)} new"
    )
    reports |= listed

    best: dict[tuple[int, int], Report] = {}
    for report in sorted(reports):
        best.setdefault((report.fiscal_year, report.fiscal_quarter), report)
    return sorted(best.values()), captures


# --- parsing ------------------------------------------------------------------


def parse_value(text: str) -> int | None:
    """One cell: a count, 0 for "-" (represents zero), None for "D" (withheld) and blanks."""
    text = text.strip().replace(",", "")
    return int(text) if text.isdigit() else 0 if text == "-" else None


def parse_values(cells: list[str]) -> Values | None:
    """The twelve counts of a row, or None when it is not a data row."""
    cells = [cell.strip() for cell in cells]
    if (
        len(cells) != 12
        or not any(cells)
        or not all(VALUE_TOKEN.match(c) for c in cells)
    ):
        return None
    return [parse_value(cell) for cell in cells]


def state_name(text: str) -> str | None:
    return STATES_BY_NORMALIZED_NAME.get(re.sub(r"[^a-z]", "", text.lower()))


def is_total_label(text: str) -> bool:
    return re.sub(r"[^a-z]", "", text.lower()) in ("total", "grandtotal")


def parse_grid(grid: list[list[str]]) -> ParsedReport:
    """Parse the CSV/XLSX layout: label columns, then 3 x (Received, Approved, Denied, Pending)."""
    value_columns: list[int] | None = None
    for row in grid:
        received = [i for i, cell in enumerate(row) if "received" in cell.lower()]
        if len(received) == 3:
            value_columns = [i + k for i in received for k in range(4)]
            break
    if value_columns is None:
        raise ValueError("no header row with three 'Received' columns")

    offices: list[OfficeRow] = []
    totals: list[Values] = []
    state: str | None = None
    for row in grid:
        cells = [cell.strip() for cell in row] + [""] * (
            max(value_columns) + 1 - len(row)
        )
        labels = cells[: value_columns[0]]
        code = next((c for c in labels if CODE.match(c)), None)
        names = [" ".join(c.split()) for c in labels if c and c != code]
        if not names:
            continue
        name = names[-1]
        values = parse_values([cells[i] for i in value_columns])
        if is_total_label(name) and values is not None:
            totals.append(values)
        elif values is None:
            # A heading: a state, a country in the international offices
            # section, or a label like "Field Office by State". Only checked
            # on rows without counts, as the Washington field office (in DC)
            # shares its name with the state.
            if labels[0]:
                state = state_name(name)
            continue
        elif code is not None:
            offices.append(OfficeRow(state, name, code, values))
        elif labels[0] or state is None:
            # No office code and either the name sits in the state column or it
            # is under a country heading: an international office (code "N/A"
            # in later reports) or a stray line, not a field office.
            if labels[0] and state_name(name) is not None:
                state = state_name(name)
            continue
        else:
            offices.append(OfficeRow(state, name, code, values))
    return ParsedReport(offices, totals)


def parse_csv(content: bytes) -> ParsedReport:
    text = content.decode("utf-8-sig", errors="replace")
    return parse_grid(list(csv.reader(io.StringIO(text))))


def parse_xlsx(content: bytes) -> ParsedReport:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    grid = [
        ["" if cell is None else str(cell) for cell in row]
        for row in workbook.worksheets[0].iter_rows(values_only=True)
    ]
    return parse_grid(grid)


def parse_pdf(content: bytes) -> ParsedReport:
    """Parse the PDF layout from its text layer: one office per line, followed by its 12 counts."""
    offices: list[OfficeRow] = []
    totals: list[Values] = []
    state: str | None = None
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        lines = [
            line.strip()
            for page in pdf.pages
            for line in (page.extract_text() or "").splitlines()
        ]
    for line in lines:
        if (known_state := state_name(line)) is not None:
            state = known_state
        elif (match := PDF_TOTAL_LINE.match(line)) is not None:
            # Totals often come out with digits split off ("1 94,819"); only
            # keep the line when it is exactly twelve well-formed numbers.
            values = parse_values(match["values"].split())
            if values is not None:
                totals.append(values)
        elif (match := PDF_OFFICE_LINE.match(line)) is not None:
            values = parse_values(match["values"].split())
            if values is not None:
                offices.append(OfficeRow(state, match["name"], match["code"], values))
    return ParsedReport(offices, totals)


def parse_report(report: Report, content: bytes) -> ParsedReport:
    parser = {"xlsx": parse_xlsx, "csv": parse_csv, "pdf": parse_pdf}[report.ext]
    parsed = parser(content)
    if len(parsed.offices) < 60:
        raise ValueError(
            f"only {len(parsed.offices)} field offices parsed from {report.url}; the layout must have changed"
        )
    return parsed


# --- assembly -----------------------------------------------------------------


def normalize_name(name: str) -> str:
    name = re.sub(r"[^a-z ]", "", name.lower())
    name = re.sub(r"\b(saint|st)\b", "st", name)
    name = re.sub(r"\b(fort|ft)\b", "ft", name)
    return re.sub(r"\s+", " ", name).strip()


def slugify(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")


def totals_of(values: Values) -> dict[str, int | None]:
    """The combined (non-military + military) received/approved/denied/pending counts of a row."""
    received, approved, denied, pending = values[8:12]
    return {
        "received": received,
        "approved": approved,
        "denied": denied,
        "pending": pending,
    }


def sum_of(rows: list[OfficeRow]) -> dict[str, int | None]:
    return {
        key: sum(totals_of(row.counts)[key] or 0 for row in rows)
        for key in ("received", "approved", "denied", "pending")
    }


def build_dataset(parsed_reports: list[tuple[Report, ParsedReport]]) -> dict[str, Any]:
    # Reports before FY2017 name the offices without their codes; resolve those
    # names against every report that has both, most specific match first.
    codes_by_name: dict[tuple[str | None, str], str] = {}
    for _, parsed in parsed_reports:
        for row in parsed.offices:
            if row.code is not None:
                codes_by_name.setdefault(
                    (row.state, normalize_name(row.name)), row.code
                )
                codes_by_name.setdefault((None, normalize_name(row.name)), row.code)

    periods = []
    offices: dict[str, dict[str, Any]] = {}
    total: dict[str, dict[str, int | None]] = {}
    for report, parsed in parsed_reports:
        key, start, end = report.calendar_quarter
        periods.append(
            {
                "quarter": key,
                "start": start.isoformat(),
                "end": end.isoformat(),
                "fiscalYear": report.fiscal_year,
                "fiscalQuarter": report.fiscal_quarter,
                "source": report.url,
            }
        )
        rows: list[OfficeRow] = []
        for row in parsed.offices:
            code = (
                row.code
                or codes_by_name.get((row.state, normalize_name(row.name)))
                or codes_by_name.get((None, normalize_name(row.name)))
            )
            if code is None:
                print(
                    f"::warning::{key}: no office code known for {row.name!r} ({row.state}); skipped"
                )
                continue
            rows.append(OfficeRow(row.state, row.name, code, row.counts))
            # The newest report a code appears in names it; reports are in chronological order.
            office = offices.setdefault(code, {"code": code, "quarters": {}})
            office["name"] = NAME_FIXES.get(code, " ".join(row.name.split()))
            office["state"] = row.state
            office["quarters"][key] = totals_of(row.counts)
        total[key] = (
            totals_of(parsed.total) if parsed.total is not None else sum_of(rows)
        )

    for office in offices.values():
        state_code = STATES.get(office["state"]) if office["state"] else None
        office["stateCode"] = state_code
        office["slug"] = slugify(f"{office['name']} {state_code or ''}")
    slugs = [office["slug"] for office in offices.values()]
    if len(set(slugs)) != len(slugs):
        raise ValueError(
            f"duplicate office slugs: {sorted(s for s in slugs if slugs.count(s) > 1)}"
        )

    return {
        "periods": periods,
        "offices": sorted(offices.values(), key=lambda office: office["code"]),
        "total": total,
    }


def main() -> int:
    reports, captures = discover_reports()
    print(
        f"{len(reports)} quarters, {reports[0].fiscal_year} Q{reports[0].fiscal_quarter} to {reports[-1].fiscal_year} Q{reports[-1].fiscal_quarter}"
    )
    parsed_reports = []
    for report in reports:
        parsed = parse_report(report, download_report(report, captures))
        parsed_reports.append((report, parsed))
        print(
            f"FY{report.fiscal_year} Q{report.fiscal_quarter} ({report.ext}): {len(parsed.offices)} offices,"
            f" total {'from report' if parsed.total else 'summed'}"
        )
    dataset = build_dataset(parsed_reports)
    OUTPUT_PATH.write_text(json.dumps(dataset, indent=2, sort_keys=True) + "\n")
    print(
        f"wrote {OUTPUT_PATH} with {len(dataset['offices'])} offices over {len(dataset['periods'])} quarters"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
